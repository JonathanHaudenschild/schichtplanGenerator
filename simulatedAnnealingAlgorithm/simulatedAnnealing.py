import random
import math
import numpy as np
import csv
import openpyxl
from openpyxl.styles import PatternFill, Font
from colorhash import ColorHash

index_name_list = [
    (0, 'Berdi'),
    (1, 'Alejandro'),
    (2, 'Alexander Alex / Zwalle'),
    (3, 'Alina'),
    (4, 'Anna Ansen'),
    (5, 'Annabelle'),
    (6, 'Antonia'),
    (7, 'Bendix'),
    (8, 'Benjamin Ben'),
    (9, 'Bruno'),
    (10, 'Christian Chre'),
    (11, 'Christin Tine'),
    (12, 'Christoph Chris'),
    (13, 'Daniel'),
    (14, 'Elisa Lush'),
    (15, 'Fabs Fabs'),
    (16, 'Frederik Freddy'),
    (17, 'Gerrit'),
    (18, 'Geza'),
    (19, 'Grace'),
    (20, 'Hanna'),
    (21, 'Jannik'),
    (22, 'Jasper'),
    (23, 'Jonas'),
    (24, 'Judith'),
    (25, 'Lea'),
    (26, 'Leonie Leo'),
    (27, 'Lorenz'),
    (28, 'Lydia Lulu'),
    (29, 'Margret'),
    (30, 'Marie-Sophie Mary'),
    (31, 'Marlene'),
    (32, 'Max B'),
    (33, 'Maximilian Murmelo'),
    (34, 'Mohamed'),
    (35, 'Pablo'),
    (36, 'Paul'),
    (37, 'Philipp Phil'),
    (38, 'Philipp'),
    (39, 'Rani'),
    (40, 'Robert Bört'),
    (41, 'Sophie'),
    (42, 'Stefanie Steffi'),
    (43, 'Torben'),
    (44, 'Undine'),
    (45, 'Christian'),
    (46, 'Fabio'),
    (47, 'Iinaroosa Lina'),
    (48, 'Ksenia'),
    (49, 'Lisa'),
    (50, 'Mona'),
    (51, 'Natalie'),
    (52, 'Nathie'),
    (53, 'Nico'),
    (54, 'Paul B'),
    (55, 'Philipp Phipsi'),
    (56, 'Rose'),
    (57, 'Sasika'),
    (58, 'Platzperson'),
    (59, 'Platzperson 1'),
    (60, 'Platzperson 2'),
    (61, 'Platzperson 3'),
    (62, 'Platzperson 4'),
    (63, 'Jonny'),
    (64, 'JULIUS'),
    (65, 'JONATHAN'),
    (66, 'MAX'),
    (67, 'ANNA B'),
    (68, 'SEBI'),
    (69, 'ANNEKE'),
    (70, 'MITCH'),
    (71, 'DANA'),
    (72, 'ALI'),
    (73, 'MARIA'),
    (74, 'FABI'),
    (75, 'LUKAS')
]

shift_name_list = [
    (0, '01:00 - 07:00'),
    (1, '07:00 - 13:00'),
    (2, '13:00 - 19:00'),
    (3, '19:00 - 01:00'),
]
# Update the preference matrix based on the preferences
# Positive values indicate a preference to work together, and negative values indicate a preference not to work together
preference_list = [
    (3, 17, 5),  # Person 0 and person 1 have a preference of 5 to work together
    (34, 70, -7), # Person 2 and person 3 have a preference of -3 (do not want to work together)
    (45, 50, 5),
    (36, 28, 5),
    (6, 11, 1),
    (6, 46, 1),
    (11, 46, 1),
    # Add more preferences here
]

preferred_shift_type_list = [
    (4, 1),  # Person 3 prefers to work check-in shifts
    (34, 1)
]


preferred_shift_list = [
    (0, (-3, -1, -2, 0)),
    (1, (-3, -1, -2, 0)),
    (2, (-3, -1, -2, 0)),
    (3, (-3, -1, -2, 0)),
    (4, (-3, -1, -2, 0)),
]
    

# Update the experience matrix based on the experience data
# neu: 0
# 1 Jahr: +1
# >1 Jahr: +2
experience_list = [
    (0, 2),  # Person 0 has experience of 5
    (1, 2),  # Person 1 has experience of 4
    (1, 2),  # Person 1 has experience of 4
    # Add more experience here
]

# Update the unavailability matrix based on the unavailability data
unavailability_list = [
    (0, (2, 3)),  # Person 0 is unavailable for shifts 2
    (4, (2, 3)),  # Person 0 is unavailable for shift 2
    (1, (2, 3)),  # Person 1 is unavailable for shift 4
    (16, (0, 1, 2, 3, 4, 5, 6, 7)),  # Person 1 is unavailable for shift 4
    # Add more unavailability data here
]

# Update the off-day matrix based on the off-day data
offDay_list= [
    (0, (2, 3)),  # Person 0 is unavailable for shift 2
    (1, (3, 5)),  # Person 1 is unavailable for shift 4
    # Add more unavailability data here
]

shift_type_list= [
    (0, 1),  # Shift 0 is a check-in shift
    (1, 1),  # Shift 1 is a check-in shift
    (2, 1),  # Shift 2 is a check-in shift
    (3, 1),  # Shift 3 is a check-in shift
    (4, 1),  # Shift 4 is a check-in shift
]

# Update the off-day matrix based on the off-day data
mandatory_shifts_list= [
    (8, 9, 10, 11),
    (12, 13, 14, 15),
    (16, 17, 18, 19, 20)
]



shift_capacity_list= [
    # 28.06.2023
    (0, (16, 17)),
    (1, (18, 19)),
    (2, (17, 18)),
    (3, (16, 17)),
    # 29.06.2023
    (4, (12, 14)),
    (5, (15, 17)),
    (6, (15, 17)),
    (7, (14, 16)),
    # 30.06.2023
    (8, (10, 12)),
    (9, (13, 16)),
    (10, (13, 16)),
    (11, (12, 14)),
    # 01.07.2023
    (12, (8, 10)),
    (13, (9, 11)),
    (14, (9, 11)),
    (15, (9, 11)),
    # 02.07.2023
    (16, (7, 10)),
    (17, (8, 10)),
    (18, (8, 10)),
    (19, (8, 10)),
    # 03.07.2023
    (20, (6, 7)),
    (21,(5, 5)),
]

person_capacity_list   = [
    (64, 1),   # Person 2 has a capacity of 4
    (65, 1),   # Person 3 has a capacity of 4
    (66, 1),   # Person 4 has a capacity of 4
    (67, 1),   # Person 5 has a capacity of 4
    (68, 1),   # Person 6 has a capacity of 4
    (69, 1),   # Person 7 has a capacity of 4
    (70, 1),   # Person 8 has a capacity of 4
    (71, 1),   # Person 9 has a capacity of 4
    (72, 1),   # Person 10 has a capacity of 4
    (73, 1),   # Person 11 has a capacity of 4
    (74, 1),   # Person 12 has a capacity of 4
    (75, 1),   # Person 13 has a capacity of 4
]


num_of_shifts = len(shift_capacity_list)  # number of shifts
num_people = len(index_name_list)   # number of people
num_of_shifts_per_person = 4  # number of shifts per person
initial_temperature = 1000
cooling_rate = 0.9999

def generate_initial_solution(x, y):
    solution = [set() for _ in range(x)]
    

    for person in range(y):
        assigned_shifts = set()
        while len(assigned_shifts) < create_persons_capacity_array(person_capacity_list)[person]:
            shift = random.randint(0, x - 1)
            if len(solution[shift]) < create_shift_capacity_matrix(shift_capacity_list)[1][shift] and shift not in assigned_shifts:
                # Higher probability of adding a person if the shift has less than the minimum people
                if len(solution[shift]) < create_shift_capacity_matrix(shift_capacity_list)[0][shift] or random.random() < 0.30:
                    solution[shift].add(person)
                    assigned_shifts.add(shift)
    return solution


def get_neighbor(solution, unavailability_matrix, max_attempts=1000):
    attempts = 0
    while attempts < max_attempts:
        index1 = random.randint(0, len(solution) - 1)
        index2 = random.randint(0, len(solution) - 1)
        shift1 = solution[index1].copy()
        shift2 = solution[index2].copy()

        if len(shift1) > create_shift_capacity_matrix(shift_capacity_list)[0][index1] and len(shift2) < create_shift_capacity_matrix(shift_capacity_list)[0][index2] :
          
            person_to_move = get_random_element(shift1)

            if person_to_move not in shift2 and consecutive_shifts(solution, index2, person_to_move) == 0 and unavailability(index2, unavailability_matrix, person_to_move) == 0:
                shift1.remove(person_to_move)
                shift2.add(person_to_move)
                new_solution = solution.copy()
                new_solution[index1] = shift1
                new_solution[index2] = shift2
                return new_solution  # The neighbor solution satisfies both hard constraints
        else:
            # Swap people between the shifts if possible
            person1 = get_random_element(shift1)
            person2 = get_random_element(shift2)

            if person1 not in shift2 and person2 not in shift1 and consecutive_shifts(solution, index2, person1) == 0 and consecutive_shifts(solution, index1, person2) == 0 and unavailability(index2, unavailability_matrix, person1) == 0 and unavailability(index1, unavailability_matrix, person2) == 0:
                shift1.remove(person1)
                shift1.add(person2)
                shift2.remove(person2)
                shift2.add(person1)
                new_solution = solution.copy()
                new_solution[index1] = shift1
                new_solution[index2] = shift2
                return new_solution  # The neighbor solution satisfies both hard constraints

        attempts += 1

    # Return the original solution if no valid neighbor is found after max_attempts
    print("No valid neighbor found after", max_attempts, "attempts")
    return solution



def get_random_element(s):
    return random.choice(list(s))


def adaptive_cooling_rate(initial_temperature, min_temperature, cooling_factor):
    return max(min_temperature, initial_temperature * cooling_factor)


def acceptance_probability(old_cost, new_cost, temperature):
    if new_cost < old_cost:
        return 1
    else:
        return math.exp(-(abs(new_cost - old_cost) / temperature))

def simulated_annealing(x, y, initial_temperature, cooling_rate, max_iterations_without_improvement=1000):
    current_solution = generate_initial_solution(x, y)
    current_cost = cost_function(current_solution)
    print(f"Current solution: {current_solution}")
    print(f"Current cost: {current_cost}")
    temperature = initial_temperature
    iterations_without_improvement = 0

    while temperature > 1 and iterations_without_improvement < max_iterations_without_improvement:
        new_solution = get_neighbor(current_solution, create_unavailability_matrix(unavailability_list))
        new_cost = cost_function(new_solution)
        if acceptance_probability(current_cost, new_cost, temperature) > random.random():
            current_solution = new_solution
            current_cost = new_cost
            iterations_without_improvement = 0  # Reset the counter when an improvement is found
        else:
            iterations_without_improvement += 1

        temperature *= cooling_rate

    return current_solution, current_cost



def create_preference_matrix(preference_list):
    preference_matrix = [[0 for _ in range(num_people)] for _ in range(num_people)]
    for person1, person2, preference in preference_list:
        preference_matrix[person1][person2] = preference
        preference_matrix[person2][person1] = preference
    return preference_matrix

def create_persons_capacity_array(capacity_list):
    capacity_array = [num_of_shifts_per_person] * num_people
    for capacity in capacity_list:
        capacity_array[capacity[0]] = capacity[1]
    return capacity_array

def create_preferred_shift_type_array(pref_shift_type_list):
    pref_shift_type_array = [0] * num_people
    for pref_shift_type in pref_shift_type_list:
        pref_shift_type_array[pref_shift_type[0]] = pref_shift_type[1]
    return pref_shift_type_array

def create_experience_array(experience_list):
    experience_array = [1] * num_people
    for experience in experience_list:
        experience_array[experience[0]] = experience[1]
    return experience_array

def create_shift_type_array(shift_type_list):
    shift_type_array = [0] * num_of_shifts
    for shift_type in shift_type_list:
        shift_type_array[shift_type[0]] = shift_type[1]
    return shift_type_array


def create_preferred_shift_matrix(pref_shift_list):
    matrix = [[0] * 4 for _ in range(num_people)]
    for i, (_, shift_values) in enumerate(pref_shift_list):
        matrix[i] = list(shift_values)
    return matrix

def create_unavailability_matrix(unavailability_list):
    unavailability_matrix = [[0 for _ in range(num_of_shifts)] for _ in range(num_people)]
    for person, shifts in unavailability_list:
        for shift in shifts:
            unavailability_matrix[person][shift] = 1
    return unavailability_matrix

def create_shift_capacity_matrix(shift_capacity_list):
    shift_capacity_matrix = [[0 for _ in range(num_of_shifts)] for _ in range(2)]
    for shift, capacity in shift_capacity_list:
        shift_capacity_matrix[0][shift] = capacity[0]
        shift_capacity_matrix[1][shift] = capacity[1]
    return shift_capacity_matrix


def cost_function(solution):
    # # Calculate preference-based cost
    pref_cost = preference_cost(solution, create_preference_matrix(preference_list))
    exp_cost = mixedExperience_cost(solution, create_experience_array(experience_list))

    cons_cost = consecutive_shifts_cost(solution)
    # # Calculate off-day-based cost
    oday_cost = offDay_cost(solution, create_unavailability_matrix(offDay_list))
    # unavail_cost = unavailability_cost(solution, create_unavailability_matrix(unavailability_list))
    man_cost = mandatory_shifts_cost(solution, mandatory_shifts_list)

    tim_cost = time_between_cost(solution)

    pref_shift_cost = preferred_shift_cost(solution, create_preferred_shift_matrix(preferred_shift_list))
    # shift_dist_cost = shift_distribution_cost(solution)
    shift_type_cost = shift_type_com_cost(solution, shift_type_list, preferred_shift_type_list)
    # Add other cost components if necessary
    total_cost = pref_cost + exp_cost + oday_cost + pref_shift_cost + cons_cost + shift_type_cost + tim_cost
    return total_cost 

def shift_type_com_cost(solution, shift_type_list, pref_shift_type_list):
    total_cost = 0
    for shift_index, shift in enumerate(solution):
        for person in shift:
            if create_shift_type_array(shift_type_list)[shift_index] != create_preferred_shift_type_array(pref_shift_type_list)[person]:
                total_cost += 3
    return total_cost

def mandatory_shifts_cost(solution, mandatory_shifts_list):
    total_cost = 0
    for person in range(num_people):
        person_shift_count = 0
        for shift_index, shift in enumerate(solution):
            if person in shift:
                for mandatory_shifts in mandatory_shifts_list:
                    if shift_index in mandatory_shifts:
                        person_shift_count += 1
                        break  # Exit the inner for-loop once a mandatory shift is found
        if person_shift_count < 2:
            total_cost += 1
        if person_shift_count > 2:
            total_cost += 2
    return total_cost

def time_between_cost(solution):
    total_cost = 0
    for person in range(num_people):
        old_index = 0
        for shift_index, shift in enumerate(solution):
            if person in shift:
                if old_index != 0:
                    if shift_index - old_index == 4:
                        total_cost -= 3
                    old_index = shift_index
    return total_cost


def mixedExperience_cost(solution, experience_array):
    total_cost = 0
    for shift_index, shift in enumerate(solution):
        total_experience = 0
        for person in shift:
            total_experience += experience_array[person]
        if total_experience < 5:
            total_cost += 5
    return total_cost 

def preferred_shift_cost(solution, preferred_shift_matrix):
    total_cost = 0
    for shift_index, shift in enumerate(solution):
        for person in shift:
            total_cost += preferred_shift_matrix[person][shift_index % 4] 
    return total_cost
                
def shift_distribution_cost(solution):
    total_cost = 0
    shift_types = [set() for _ in range(num_people)]
    
    for shift_index, shift in enumerate(solution):
        for person in shift:
            shift_type = shift_index % 4
            if shift_type not in shift_types[person]:
                shift_types[person].add(shift_type)
            else:
                total_cost += 2
                
    return total_cost


def offDay_cost(solution, unavailability_matrix):
    total_cost = 0
    for shift_index, shift in enumerate(solution):
        for person in shift:
            if unavailability_matrix[person][shift_index]:
                total_cost += 6  # Penalize the cases when a person is assigned to an unavailable shift
    return total_cost

def unavailability_cost(solution, unavailability_matrix):
    total_cost = 0
    for shift_index, shift in enumerate(solution):
        for person in shift:
            if unavailability_matrix[person][shift_index]:
                total_cost += 10  # Penalize the cases when a person is assigned to an unavailable shift
    return total_cost

def unavailability(shift_index, unavailability_matrix, person):
        if unavailability_matrix[person][shift_index]:
            return 1  # Penalize the cases when a person is assigned to an unavailable shift
        return 0

def preference_cost(solution, preference_matrix):
    total_cost = 0
    for shift in solution:
        for person1 in shift:
            for person2 in shift:
                if person1 != person2:
                    total_cost -= preference_matrix[person1][person2]
    return total_cost

def are_shifts_filled_evenly_cost(solution):
    shift_proportions = []
    total_cost = 0
    for shift_index, shift in enumerate(solution):
        assigned_people = len(shift)
        max_capacity = create_shift_capacity_matrix(shift_capacity_list)[1][shift_index]
        proportion = assigned_people / max_capacity
        shift_proportions.append(proportion)

    min_proportion = min(shift_proportions)
    max_proportion = max(shift_proportions)

    # You can adjust the threshold value (e.g., 0.1) based on the desired level of evenness
    if max_proportion - min_proportion > 0.2:
        total_cost += 10

    return total_cost

def consecutive_shifts_cost(solution):
    total_cost = 0
    for shift_index, shift in enumerate(solution):
        for person in shift:
            total_cost += consecutive_shifts(solution, shift_index, person)
    return total_cost

def consecutive_shifts(solution, shift_index, person):
    # Check previous shifts
    for i in range(1, 3):
        if shift_index - i >= 0:
            previous_shift = solution[shift_index - i]
            if person in previous_shift:
                return 1  # Return the penalty as soon as the constraint is violated

    # Check next shifts
    for i in range(1, 3):
        if shift_index + i < len(solution):
            next_shift = solution[shift_index + i]
            if person in next_shift:
                return 1

    return 0  # No constraint violations were found

def replace_numbers_with_names(solution, index_name_list):
    name_solution = []
    for shift in solution:
        name_shift = set()
        for number in shift:
            name = next(name for index, name in index_name_list if index == number)
            name_shift.add(name)
        name_solution.append(name_shift)
    return name_solution

if __name__ == "__main__":
    best_solution, best_cost = simulated_annealing(num_of_shifts, num_people, initial_temperature, cooling_rate, max_iterations_without_improvement=1000) 
    best_solution_with_names = replace_numbers_with_names(best_solution, index_name_list) 
    print(f"Best solution: {best_solution}")
    print(f"Best solution with names: {best_solution_with_names}")
    print(f"Best cost: {best_cost}")
    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write shift names as headers
    for col_index, _ in enumerate(best_solution_with_names):
        shift_name = shift_name_list[col_index % len(shift_name_list)][1]
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = shift_name


        # Write the data to the worksheet and apply a unique color to each name
    name_colors = {}
    white_font = Font(color='FFFFFF')  # Set font color to white
    for col_index, shift in enumerate(best_solution_with_names, start=1):
        for row_index, name in enumerate(shift, start=2):
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = name

            # Generate a unique color for each name if not already generated
            if name not in name_colors:
                color = ColorHash(name).hex
                color = 'FF' + color[1:]  # Add alpha channel to the hex color
                name_colors[name] = PatternFill(start_color=color, end_color=color, fill_type='solid')

            # Apply the unique color to the cell
            cell.fill = name_colors[name]
            cell.font = white_font

    # Save the workbook as an Excel file
    workbook.save('shifts_with_unique_colors.xlsx')